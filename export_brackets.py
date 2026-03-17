"""
Export 20 brackets to Excel spreadsheet.
Each bracket gets its own sheet with every game clearly listed.
Also creates a summary sheet with advancement probabilities.
"""

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

import csv
import os
from collections import Counter
from team_data import TEAMS, BRACKET_MATCHUPS, ESPN_SCORING
from simulator import count_upsets_in_bracket
from algorithm import predict_game_total


# Color scheme
COLORS = {
    "header_fill": "1B3A5C",
    "header_font": "FFFFFF",
    "round_fill": "37474F",
    "region_fill": "455A64",
    "seed_1_fill": "C5E1A5",
    "seed_2_3_fill": "E8F5E9",
    "seed_4_6_fill": "FFF9C4",
    "seed_7_9_fill": "FFE0B2",
    "seed_10_12_fill": "FFCCBC",
    "seed_13_16_fill": "EF9A9A",
    "upset_fill": "FF6F00",
    "upset_font": "FFFFFF",
    "champion_fill": "FFD700",
    "ff_fill": "CE93D8",
    "border_color": "BDBDBD",
    "alt_row": "F5F5F5",
}

THIN_BORDER = Border(
    left=Side(style='thin', color=COLORS["border_color"]),
    right=Side(style='thin', color=COLORS["border_color"]),
    top=Side(style='thin', color=COLORS["border_color"]),
    bottom=Side(style='thin', color=COLORS["border_color"]),
)


def get_seed_fill(seed):
    if seed == 1:
        return COLORS["seed_1_fill"]
    elif seed <= 3:
        return COLORS["seed_2_3_fill"]
    elif seed <= 6:
        return COLORS["seed_4_6_fill"]
    elif seed <= 9:
        return COLORS["seed_7_9_fill"]
    elif seed <= 12:
        return COLORS["seed_10_12_fill"]
    else:
        return COLORS["seed_13_16_fill"]


def styled_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=True):
    """Helper to write a styled cell."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = THIN_BORDER
    return cell


def write_header_row(ws, row, col_start, col_end, text, merge=True):
    """Write a dark header bar spanning columns."""
    if merge and col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = Font(bold=True, size=12, color=COLORS["header_font"])
    cell.fill = PatternFill("solid", fgColor=COLORS["header_fill"])
    cell.alignment = Alignment(horizontal="center")
    return cell


# =============================================================================
# INDIVIDUAL BRACKET SHEET — clean table layout showing all 63 games
# =============================================================================

def write_bracket_sheet(wb, bracket):
    """Write one bracket as a clean table with every game listed."""
    num = bracket.get("_number", 0)
    strategy = bracket.get("_strategy", "")
    champion = bracket["Champion"]
    champ_seed = TEAMS[champion]["seed"]
    upset_count = count_upsets_in_bracket(bracket)

    ws = wb.create_sheet(f"Bracket {num}")

    # Column layout:
    # A: Higher Seed Team  B: Lower Seed Team  C: Winner  D: Win Prob  E: Upset?
    col_widths = {"A": 26, "B": 26, "C": 26, "D": 10, "E": 8}
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width

    row = 1

    # Title
    write_header_row(ws, row, 1, 5, f"BRACKET #{num}: {strategy}")
    row += 1

    # Champion banner
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = ws.cell(row=row, column=1, value=f"CHAMPION: ({champ_seed}) {champion}  |  Upsets: {upset_count}")
    cell.font = Font(bold=True, size=14)
    cell.fill = PatternFill("solid", fgColor=COLORS["champion_fill"])
    cell.alignment = Alignment(horizontal="center")
    row += 2

    # ---- Process each region ----
    regions = ["East", "West", "South", "Midwest"]
    for region in regions:
        results = bracket[region]
        matchups = BRACKET_MATCHUPS[region]

        # Region header
        write_header_row(ws, row, 1, 5, f"{region.upper()} REGION")
        row += 1

        # ---- ROUND OF 64 ----
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell = ws.cell(row=row, column=1, value="Round of 64")
        cell.font = Font(bold=True, size=10, color=COLORS["header_font"])
        cell.fill = PatternFill("solid", fgColor=COLORS["round_fill"])
        row += 1

        # Column headers
        for col, label in enumerate(["Team A", "Team B", "Winner", "", ""], 1):
            if label:
                cell = ws.cell(row=row, column=col, value=label)
                cell.font = Font(bold=True, size=9, color="666666")
                cell.alignment = Alignment(horizontal="center")
        row += 1

        for i, (team_a, team_b) in enumerate(matchups):
            winner = results["R64"][i]
            seed_a = TEAMS[team_a]["seed"]
            seed_b = TEAMS[team_b]["seed"]
            seed_w = TEAMS[winner]["seed"]
            loser = team_b if winner == team_a else team_a
            seed_l = TEAMS[loser]["seed"]
            is_upset = seed_w > seed_l

            alt = COLORS["alt_row"] if i % 2 == 1 else None

            styled_cell(ws, row, 1, f"({seed_a}) {team_a}",
                        fill=get_seed_fill(seed_a),
                        font=Font(bold=(team_a == winner)))
            styled_cell(ws, row, 2, f"({seed_b}) {team_b}",
                        fill=get_seed_fill(seed_b),
                        font=Font(bold=(team_b == winner)))

            if is_upset:
                styled_cell(ws, row, 3, f"({seed_w}) {winner}",
                            fill=COLORS["upset_fill"],
                            font=Font(bold=True, color=COLORS["upset_font"]))
                styled_cell(ws, row, 4, "UPSET", font=Font(bold=True, color="FF6F00", size=9),
                            alignment=Alignment(horizontal="center"))
            else:
                styled_cell(ws, row, 3, f"({seed_w}) {winner}",
                            fill=get_seed_fill(seed_w),
                            font=Font(bold=True))

            row += 1

        row += 1  # Spacing

        # ---- ROUND OF 32 ----
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell = ws.cell(row=row, column=1, value="Round of 32")
        cell.font = Font(bold=True, size=10, color=COLORS["header_font"])
        cell.fill = PatternFill("solid", fgColor=COLORS["round_fill"])
        row += 1

        for col, label in enumerate(["Team A", "Team B", "Winner", "", ""], 1):
            if label:
                cell = ws.cell(row=row, column=col, value=label)
                cell.font = Font(bold=True, size=9, color="666666")
                cell.alignment = Alignment(horizontal="center")
        row += 1

        for i in range(4):
            team_a = results["R64"][i * 2]
            team_b = results["R64"][i * 2 + 1]
            winner = results["R32"][i]
            seed_a = TEAMS[team_a]["seed"]
            seed_b = TEAMS[team_b]["seed"]
            seed_w = TEAMS[winner]["seed"]
            loser = team_b if winner == team_a else team_a
            seed_l = TEAMS[loser]["seed"]
            is_upset = seed_w > seed_l

            styled_cell(ws, row, 1, f"({seed_a}) {team_a}",
                        fill=get_seed_fill(seed_a),
                        font=Font(bold=(team_a == winner)))
            styled_cell(ws, row, 2, f"({seed_b}) {team_b}",
                        fill=get_seed_fill(seed_b),
                        font=Font(bold=(team_b == winner)))

            if is_upset:
                styled_cell(ws, row, 3, f"({seed_w}) {winner}",
                            fill=COLORS["upset_fill"],
                            font=Font(bold=True, color=COLORS["upset_font"]))
                styled_cell(ws, row, 4, "UPSET", font=Font(bold=True, color="FF6F00", size=9),
                            alignment=Alignment(horizontal="center"))
            else:
                styled_cell(ws, row, 3, f"({seed_w}) {winner}",
                            fill=get_seed_fill(seed_w),
                            font=Font(bold=True))

            row += 1

        row += 1  # Spacing

        # ---- SWEET 16 ----
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell = ws.cell(row=row, column=1, value="Sweet 16")
        cell.font = Font(bold=True, size=10, color=COLORS["header_font"])
        cell.fill = PatternFill("solid", fgColor=COLORS["round_fill"])
        row += 1

        for col, label in enumerate(["Team A", "Team B", "Winner", "", ""], 1):
            if label:
                cell = ws.cell(row=row, column=col, value=label)
                cell.font = Font(bold=True, size=9, color="666666")
                cell.alignment = Alignment(horizontal="center")
        row += 1

        for i in range(2):
            team_a = results["R32"][i * 2]
            team_b = results["R32"][i * 2 + 1]
            winner = results["S16"][i]
            seed_a = TEAMS[team_a]["seed"]
            seed_b = TEAMS[team_b]["seed"]
            seed_w = TEAMS[winner]["seed"]
            loser = team_b if winner == team_a else team_a
            seed_l = TEAMS[loser]["seed"]
            is_upset = seed_w > seed_l

            styled_cell(ws, row, 1, f"({seed_a}) {team_a}",
                        fill=get_seed_fill(seed_a),
                        font=Font(bold=(team_a == winner)))
            styled_cell(ws, row, 2, f"({seed_b}) {team_b}",
                        fill=get_seed_fill(seed_b),
                        font=Font(bold=(team_b == winner)))

            if is_upset:
                styled_cell(ws, row, 3, f"({seed_w}) {winner}",
                            fill=COLORS["upset_fill"],
                            font=Font(bold=True, color=COLORS["upset_font"]))
                styled_cell(ws, row, 4, "UPSET", font=Font(bold=True, color="FF6F00", size=9),
                            alignment=Alignment(horizontal="center"))
            else:
                styled_cell(ws, row, 3, f"({seed_w}) {winner}",
                            fill=get_seed_fill(seed_w),
                            font=Font(bold=True))

            row += 1

        row += 1  # Spacing

        # ---- ELITE 8 ----
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell = ws.cell(row=row, column=1, value="Elite 8 — Region Championship")
        cell.font = Font(bold=True, size=10, color=COLORS["header_font"])
        cell.fill = PatternFill("solid", fgColor=COLORS["round_fill"])
        row += 1

        team_a = results["S16"][0]
        team_b = results["S16"][1]
        e8_winner = results["E8"]
        seed_a = TEAMS[team_a]["seed"]
        seed_b = TEAMS[team_b]["seed"]
        seed_w = TEAMS[e8_winner]["seed"]
        loser = team_b if e8_winner == team_a else team_a
        seed_l = TEAMS[loser]["seed"]
        is_upset = seed_w > seed_l

        styled_cell(ws, row, 1, f"({seed_a}) {team_a}",
                    fill=get_seed_fill(seed_a),
                    font=Font(bold=(team_a == e8_winner), size=11))
        styled_cell(ws, row, 2, f"({seed_b}) {team_b}",
                    fill=get_seed_fill(seed_b),
                    font=Font(bold=(team_b == e8_winner), size=11))

        styled_cell(ws, row, 3, f"({seed_w}) {e8_winner} — {region} CHAMP",
                    fill=get_seed_fill(seed_w),
                    font=Font(bold=True, size=11))
        if is_upset:
            styled_cell(ws, row, 4, "UPSET", font=Font(bold=True, color="FF6F00", size=9),
                        alignment=Alignment(horizontal="center"))

        row += 2  # Extra spacing between regions

    # ---- FINAL FOUR ----
    write_header_row(ws, row, 1, 5, "FINAL FOUR")
    row += 1

    ff = bracket["FF"]
    ff_pairings = [("South", "East"), ("Midwest", "West")]

    for i, (r_a, r_b) in enumerate(ff_pairings):
        champ_a = bracket[r_a]["E8"]
        champ_b = bracket[r_b]["E8"]
        winner = ff[i]
        seed_a = TEAMS[champ_a]["seed"]
        seed_b = TEAMS[champ_b]["seed"]
        seed_w = TEAMS[winner]["seed"]
        loser = champ_b if winner == champ_a else champ_a
        seed_l = TEAMS[loser]["seed"]
        is_upset = seed_w > seed_l

        styled_cell(ws, row, 1, f"({seed_a}) {champ_a} [{r_a}]",
                    fill=COLORS["ff_fill"],
                    font=Font(bold=(champ_a == winner), size=11))
        styled_cell(ws, row, 2, f"({seed_b}) {champ_b} [{r_b}]",
                    fill=COLORS["ff_fill"],
                    font=Font(bold=(champ_b == winner), size=11))
        styled_cell(ws, row, 3, f"({seed_w}) {winner}",
                    fill=get_seed_fill(seed_w),
                    font=Font(bold=True, size=11))
        if is_upset:
            styled_cell(ws, row, 4, "UPSET", font=Font(bold=True, color="FF6F00", size=9),
                        alignment=Alignment(horizontal="center"))

        row += 1

    row += 1

    # ---- NATIONAL CHAMPIONSHIP ----
    write_header_row(ws, row, 1, 5, "NATIONAL CHAMPIONSHIP")
    row += 1

    finalist_a = ff[0]
    finalist_b = ff[1]
    seed_fa = TEAMS[finalist_a]["seed"]
    seed_fb = TEAMS[finalist_b]["seed"]

    # Projected point total for championship game
    score_a, score_b, total = predict_game_total(finalist_a, finalist_b, is_championship=True)
    # Put winner score first
    if champion == finalist_a:
        w_score, l_score = score_a, score_b
        loser_name = finalist_b
    else:
        w_score, l_score = score_b, score_a
        loser_name = finalist_a

    styled_cell(ws, row, 1, f"({seed_fa}) {finalist_a}",
                fill=COLORS["ff_fill"],
                font=Font(bold=(finalist_a == champion), size=12))
    styled_cell(ws, row, 2, f"({seed_fb}) {finalist_b}",
                fill=COLORS["ff_fill"],
                font=Font(bold=(finalist_b == champion), size=12))
    styled_cell(ws, row, 3, f"({champ_seed}) {champion} — NATIONAL CHAMPION",
                fill=COLORS["champion_fill"],
                font=Font(bold=True, size=12))

    row += 2

    # Projected score line
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = ws.cell(row=row, column=1,
                   value=f"Projected Score: {champion} {w_score:.0f}, {loser_name} {l_score:.0f}  |  Combined Total: {total:.0f}")
    cell.font = Font(bold=True, size=11, italic=True)
    cell.alignment = Alignment(horizontal="center")

    return ws


# =============================================================================
# ALL BRACKETS COMPARISON SHEET
# =============================================================================

def write_comparison_sheet(wb, brackets):
    """Write a flat comparison of all 20 brackets side by side."""
    ws = wb.create_sheet("All Brackets Comparison")

    # Title
    last_col = 3 + len(brackets)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    write_header_row(ws, 1, 1, last_col, "All 20 Brackets — Game-by-Game Comparison")

    # Column headers
    for col, label in enumerate(["Region", "Round", "Game"], 1):
        cell = ws.cell(row=3, column=col, value=label)
        cell.font = Font(bold=True, color=COLORS["header_font"])
        cell.fill = PatternFill("solid", fgColor=COLORS["header_fill"])

    for i in range(len(brackets)):
        cell = ws.cell(row=3, column=4 + i, value=f"B{i+1}")
        cell.font = Font(bold=True, color=COLORS["header_font"])
        cell.fill = PatternFill("solid", fgColor=COLORS["header_fill"])
        cell.alignment = Alignment(horizontal="center")

    row = 4
    regions = ["East", "West", "South", "Midwest"]

    for region in regions:
        matchups = BRACKET_MATCHUPS[region]

        # R64
        for i, (team_a, team_b) in enumerate(matchups):
            seed_a = TEAMS[team_a]["seed"]
            seed_b = TEAMS[team_b]["seed"]
            ws.cell(row=row, column=1, value=region)
            ws.cell(row=row, column=2, value="R64")
            ws.cell(row=row, column=3, value=f"({seed_a}) {team_a} vs ({seed_b}) {team_b}")

            for b_idx, bracket in enumerate(brackets):
                winner = bracket[region]["R64"][i]
                loser = team_b if winner == team_a else team_a
                seed_w = TEAMS[winner]["seed"]
                is_upset = TEAMS[winner]["seed"] > TEAMS[loser]["seed"]
                cell = ws.cell(row=row, column=4 + b_idx, value=winner)
                cell.alignment = Alignment(horizontal="center")
                if is_upset:
                    cell.fill = PatternFill("solid", fgColor=COLORS["upset_fill"])
                    cell.font = Font(color=COLORS["upset_font"], bold=True, size=9)
                else:
                    cell.font = Font(size=9)
            row += 1

        # R32
        for i in range(4):
            ws.cell(row=row, column=1, value=region)
            ws.cell(row=row, column=2, value="R32")
            # Describe which R64 winners meet
            m1 = matchups[i * 2]
            m2 = matchups[i * 2 + 1]
            ws.cell(row=row, column=3,
                    value=f"R64 G{i*2+1} winner vs R64 G{i*2+2} winner")

            for b_idx, bracket in enumerate(brackets):
                winner = bracket[region]["R32"][i]
                seed_w = TEAMS[winner]["seed"]
                cell = ws.cell(row=row, column=4 + b_idx, value=f"({seed_w}) {winner}")
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(size=9, bold=True)
            row += 1

        # S16
        for i in range(2):
            ws.cell(row=row, column=1, value=region)
            ws.cell(row=row, column=2, value="S16")
            ws.cell(row=row, column=3, value=f"Sweet 16 Game {i+1}")

            for b_idx, bracket in enumerate(brackets):
                winner = bracket[region]["S16"][i]
                seed_w = TEAMS[winner]["seed"]
                cell = ws.cell(row=row, column=4 + b_idx, value=f"({seed_w}) {winner}")
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(bold=True, size=9)
                cell.fill = PatternFill("solid", fgColor=get_seed_fill(seed_w))
            row += 1

        # E8
        ws.cell(row=row, column=1, value=region)
        ws.cell(row=row, column=2, value="E8")
        ws.cell(row=row, column=3, value=f"{region} CHAMPION")
        ws.cell(row=row, column=3).font = Font(bold=True)

        for b_idx, bracket in enumerate(brackets):
            winner = bracket[region]["E8"]
            seed_w = TEAMS[winner]["seed"]
            cell = ws.cell(row=row, column=4 + b_idx, value=f"({seed_w}) {winner}")
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill("solid", fgColor=get_seed_fill(seed_w))
        row += 1
        row += 1  # Blank between regions

    # Final Four
    for i, (r_a, r_b) in enumerate([("South", "East"), ("Midwest", "West")]):
        ws.cell(row=row, column=2, value="FF")
        ws.cell(row=row, column=3, value=f"Semifinal: {r_a} vs {r_b}")
        ws.cell(row=row, column=3).font = Font(bold=True)
        for b_idx, bracket in enumerate(brackets):
            winner = bracket["FF"][i]
            seed_w = TEAMS[winner]["seed"]
            cell = ws.cell(row=row, column=4 + b_idx, value=f"({seed_w}) {winner}")
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=COLORS["ff_fill"])
        row += 1

    # Championship
    ws.cell(row=row, column=2, value="CG")
    ws.cell(row=row, column=3, value="NATIONAL CHAMPION")
    ws.cell(row=row, column=3).font = Font(bold=True, size=12)
    for b_idx, bracket in enumerate(brackets):
        champ = bracket["Champion"]
        seed_w = TEAMS[champ]["seed"]
        cell = ws.cell(row=row, column=4 + b_idx, value=f"({seed_w}) {champ}")
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill("solid", fgColor=COLORS["champion_fill"])
    row += 1

    # Projected championship total
    ws.cell(row=row, column=3, value="PROJECTED COMBINED TOTAL")
    ws.cell(row=row, column=3).font = Font(bold=True, size=10, italic=True)
    for b_idx, bracket in enumerate(brackets):
        finalist_a = bracket["FF"][0]
        finalist_b = bracket["FF"][1]
        _, _, total = predict_game_total(finalist_a, finalist_b, is_championship=True)
        cell = ws.cell(row=row, column=4 + b_idx, value=round(total))
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True, size=10)

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 40
    for i in range(len(brackets)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 18

    return ws


# =============================================================================
# SUMMARY SHEETS
# =============================================================================

def write_summary_sheet(ws, advancement_data, n_sims):
    """Write advancement probability table."""
    ws.title = "Advancement Probabilities"

    write_header_row(ws, 1, 1, 8,
                     "2026 NCAA Tournament — Model Predictions (50K Simulations)")

    headers = ["Team", "Seed", "Region", "Rd of 32", "Sweet 16",
               "Elite 8", "Final Four", "Champion"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color=COLORS["header_font"])
        cell.fill = PatternFill("solid", fgColor=COLORS["header_fill"])
        cell.alignment = Alignment(horizontal="center")

    team_probs = sorted(
        [(t, advancement_data[t].get("CHAMP", 0) / n_sims) for t in TEAMS],
        key=lambda x: x[1], reverse=True
    )

    rounds_keys = ["R32", "S16", "E8", "FF", "CHAMP"]

    for row_idx, (team, _) in enumerate(team_probs, 4):
        data = TEAMS[team]
        seed = data["seed"]
        fill = get_seed_fill(seed)

        styled_cell(ws, row_idx, 1, team, fill=fill)
        styled_cell(ws, row_idx, 2, seed, fill=fill,
                    alignment=Alignment(horizontal="center"))
        styled_cell(ws, row_idx, 3, data["region"], fill=fill)

        for col_offset, rkey in enumerate(rounds_keys):
            count = advancement_data[team].get(rkey, 0)
            pct = count / n_sims
            cell = styled_cell(ws, row_idx, 4 + col_offset, pct, fill=fill,
                               alignment=Alignment(horizontal="center"))
            cell.number_format = "0.0%"
            if pct > 0.50:
                cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 10
    for col in range(4, 9):
        ws.column_dimensions[get_column_letter(col)].width = 12


def write_champion_sheet(wb, brackets, advancement_data, n_sims):
    """Write champion distribution sheet."""
    ws = wb.create_sheet("Champion Distribution")

    write_header_row(ws, 1, 1, 5, "Championship Picks Across 20 Brackets")

    headers = ["Champion", "Seed", "Region", "# Brackets", "Model Win %"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color=COLORS["header_font"])
        cell.fill = PatternFill("solid", fgColor=COLORS["header_fill"])
        cell.alignment = Alignment(horizontal="center")

    champ_dist = Counter(b["Champion"] for b in brackets)
    for row_idx, (team, count) in enumerate(champ_dist.most_common(), 4):
        seed = TEAMS[team]["seed"]
        model_pct = advancement_data[team].get("CHAMP", 0) / n_sims
        styled_cell(ws, row_idx, 1, team, fill=get_seed_fill(seed))
        styled_cell(ws, row_idx, 2, seed, alignment=Alignment(horizontal="center"))
        styled_cell(ws, row_idx, 3, TEAMS[team]["region"])
        styled_cell(ws, row_idx, 4, count, alignment=Alignment(horizontal="center"))
        cell = styled_cell(ws, row_idx, 5, model_pct,
                           alignment=Alignment(horizontal="center"))
        cell.number_format = "0.0%"

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14


# =============================================================================
# CHAMPIONSHIP PROJECTED TOTALS MATRIX
# =============================================================================

def write_totals_matrix_sheet(wb):
    """
    Write a matrix showing projected championship game totals
    for every possible matchup among top contenders.

    Includes projected score for each team and combined total.
    Accounts for dome effect and championship game depression.
    """
    ws = wb.create_sheet("Championship Totals")

    # Top contenders most likely to reach the championship
    contenders = [
        "Duke", "Michigan", "Arizona", "Florida", "Houston",
        "Iowa State", "Purdue", "UConn", "Illinois", "Michigan State",
        "Gonzaga", "Virginia", "Arkansas", "Nebraska", "Alabama",
        "St. John's", "Vanderbilt", "Kansas",
    ]

    write_header_row(ws, 1, 1, len(contenders) + 2,
                     "Projected Championship Game Combined Totals (with dome/championship depression)")

    # Context row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(contenders) + 2)
    cell = ws.cell(row=2, column=1,
                   value="Historical avg championship total: 139.0 (2021-25)  |  "
                         "Under hits 55% of title games  |  "
                         "Dome effect: shooting drops 3-5%  |  "
                         "Lucas Oil Stadium, Indianapolis")
    cell.font = Font(italic=True, size=9, color="666666")

    # Column headers (Team B across the top)
    ws.cell(row=4, column=1, value="Team A ↓  /  Team B →").font = Font(bold=True, size=9)
    for j, team_b in enumerate(contenders):
        seed_b = TEAMS[team_b]["seed"]
        cell = ws.cell(row=4, column=j + 2, value=f"({seed_b}) {team_b}")
        cell.font = Font(bold=True, size=8, color=COLORS["header_font"])
        cell.fill = PatternFill("solid", fgColor=COLORS["header_fill"])
        cell.alignment = Alignment(horizontal="center", text_rotation=45)

    # Fill the matrix
    for i, team_a in enumerate(contenders):
        seed_a = TEAMS[team_a]["seed"]
        row = 5 + i

        # Row header
        cell = ws.cell(row=row, column=1, value=f"({seed_a}) {team_a}")
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill("solid", fgColor=get_seed_fill(seed_a))
        cell.border = THIN_BORDER

        for j, team_b in enumerate(contenders):
            col = j + 2
            if team_a == team_b:
                cell = ws.cell(row=row, column=col, value="—")
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
                cell.alignment = Alignment(horizontal="center")
                cell.border = THIN_BORDER
                continue

            score_a, score_b, total = predict_game_total(team_a, team_b, is_championship=True)
            cell = ws.cell(row=row, column=col, value=round(total))
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            cell.font = Font(size=9)

            # Color code: green=low scoring, red=high scoring
            if total <= 130:
                cell.fill = PatternFill("solid", fgColor="C8E6C9")  # Very low
            elif total <= 138:
                cell.fill = PatternFill("solid", fgColor="E8F5E9")  # Low
            elif total <= 145:
                cell.fill = PatternFill("solid", fgColor="FFF9C4")  # Medium
            elif total <= 152:
                cell.fill = PatternFill("solid", fgColor="FFE0B2")  # High
            else:
                cell.fill = PatternFill("solid", fgColor="FFCCBC")  # Very high

    # Legend
    legend_row = 5 + len(contenders) + 2
    ws.cell(row=legend_row, column=1, value="Color Legend:").font = Font(bold=True)
    legend_items = [
        ("≤130", "C8E6C9"), ("131-138", "E8F5E9"), ("139-145", "FFF9C4"),
        ("146-152", "FFE0B2"), ("153+", "FFCCBC"),
    ]
    for k, (label, color) in enumerate(legend_items):
        cell = ws.cell(row=legend_row, column=k + 2, value=label)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Detailed breakdown section below the matrix
    detail_row = legend_row + 3
    write_header_row(ws, detail_row, 1, 7,
                     "Detailed Breakdown: Most Likely Championship Matchups")
    detail_row += 1

    detail_headers = ["Matchup", "Team A Score", "Team B Score", "Combined", "Pace", "Notes"]
    for col, h in enumerate(detail_headers, 1):
        cell = ws.cell(row=detail_row, column=col, value=h)
        cell.font = Font(bold=True, color=COLORS["header_font"])
        cell.fill = PatternFill("solid", fgColor=COLORS["header_fill"])
        cell.alignment = Alignment(horizontal="center")
    detail_row += 1

    from algorithm import TEAM_PACE

    # Most likely matchups
    likely_matchups = [
        ("Duke", "Michigan", "Two highest AdjEM teams in KenPom history"),
        ("Duke", "Arizona", "Elite two-way teams; Arizona healthiest"),
        ("Michigan", "Arizona", "Top 3 defenses meeting"),
        ("Duke", "Florida", "Defending champ vs #1 overall"),
        ("Michigan", "Florida", "Best defense vs best rebounding"),
        ("Duke", "Houston", "Two elite defenses; expect grind"),
        ("Arizona", "Houston", "Houston's slow pace vs Arizona's speed"),
        ("Duke", "Iowa State", "Elite D matchup"),
        ("Michigan", "Houston", "#1 defense vs #5 defense — lowest total"),
        ("Arizona", "Florida", "Two fastest-paced contenders — highest total"),
        ("Illinois", "Purdue", "Best two offenses — most entertaining"),
        ("Duke", "UConn", "UConn cold but talent-rich"),
    ]

    for team_a, team_b, note in likely_matchups:
        score_a, score_b, total = predict_game_total(team_a, team_b, is_championship=True)
        pace_a = TEAM_PACE.get(team_a, 68.0)
        pace_b = TEAM_PACE.get(team_b, 68.0)
        avg_pace = (pace_a + pace_b) / 2
        seed_a = TEAMS[team_a]["seed"]
        seed_b = TEAMS[team_b]["seed"]

        ws.cell(row=detail_row, column=1,
                value=f"({seed_a}) {team_a} vs ({seed_b}) {team_b}")
        ws.cell(row=detail_row, column=2, value=round(score_a, 1))
        ws.cell(row=detail_row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=detail_row, column=3, value=round(score_b, 1))
        ws.cell(row=detail_row, column=3).alignment = Alignment(horizontal="center")

        total_cell = ws.cell(row=detail_row, column=4, value=round(total))
        total_cell.font = Font(bold=True, size=11)
        total_cell.alignment = Alignment(horizontal="center")

        ws.cell(row=detail_row, column=5, value=round(avg_pace, 1))
        ws.cell(row=detail_row, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=detail_row, column=6, value=note)
        ws.cell(row=detail_row, column=6).font = Font(size=9, italic=True)

        for col in range(1, 7):
            ws.cell(row=detail_row, column=col).border = THIN_BORDER

        detail_row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    for j in range(len(contenders)):
        ws.column_dimensions[get_column_letter(j + 2)].width = 10
    # Wider columns for detail section
    ws.column_dimensions[get_column_letter(1)].width = 30
    ws.column_dimensions[get_column_letter(6)].width = 45


# =============================================================================
# MAIN EXPORT
# =============================================================================

def export_to_excel(brackets, advancement_data, champion_counts, n_sims,
                    filename="output/march_madness_2026_brackets.xlsx"):
    """Export all 20 brackets and analysis to Excel."""
    if not HAS_OPENPYXL:
        print("openpyxl not installed. Falling back to CSV export.")
        export_to_csv(brackets, advancement_data, champion_counts, n_sims)
        return

    wb = openpyxl.Workbook()

    # Sheet 1: Summary
    write_summary_sheet(wb.active, advancement_data, n_sims)

    # Sheet 2: Champion distribution
    write_champion_sheet(wb, brackets, advancement_data, n_sims)

    # Sheet 3: Championship projected totals matrix
    write_totals_matrix_sheet(wb)

    # Sheets 4-23: Individual brackets
    for bracket in brackets:
        write_bracket_sheet(wb, bracket)

    # Sheet 24: All brackets comparison
    write_comparison_sheet(wb, brackets)

    os.makedirs(os.path.dirname(filename), exist_ok=True)
    wb.save(filename)
    print(f"\nSpreadsheet saved: {filename}")


def export_to_csv(brackets, advancement_data, champion_counts, n_sims):
    """Fallback CSV export if openpyxl is not installed."""
    os.makedirs("output", exist_ok=True)

    with open("output/advancement_probabilities.csv", "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Team", "Seed", "Region", "Rd of 32", "Sweet 16",
                          "Elite 8", "Final Four", "Champion"])
        team_probs = sorted(
            [(t, advancement_data[t].get("CHAMP", 0) / n_sims) for t in TEAMS],
            key=lambda x: x[1], reverse=True
        )
        for team, _ in team_probs:
            data = TEAMS[team]
            writer.writerow([
                team, data["seed"], data["region"],
                f"{advancement_data[team].get('R32', 0)/n_sims:.1%}",
                f"{advancement_data[team].get('S16', 0)/n_sims:.1%}",
                f"{advancement_data[team].get('E8', 0)/n_sims:.1%}",
                f"{advancement_data[team].get('FF', 0)/n_sims:.1%}",
                f"{advancement_data[team].get('CHAMP', 0)/n_sims:.1%}",
            ])
    print("Saved: output/advancement_probabilities.csv")

    with open("output/all_20_brackets.csv", "w", newline="") as f:
        writer = csv.writer(f)
        header = ["Region", "Round", "Game"] + [f"Bracket {i+1}" for i in range(20)]
        writer.writerow(header)

        for region in ["East", "West", "South", "Midwest"]:
            matchups = BRACKET_MATCHUPS[region]
            for i, (ta, tb) in enumerate(matchups):
                sa, sb = TEAMS[ta]["seed"], TEAMS[tb]["seed"]
                row = [region, "R64", f"({sa}) {ta} vs ({sb}) {tb}"]
                for b in brackets:
                    row.append(b[region]["R64"][i])
                writer.writerow(row)
            for i in range(4):
                row = [region, "R32", f"R32 Game {i+1}"]
                for b in brackets:
                    w = b[region]["R32"][i]
                    row.append(f"({TEAMS[w]['seed']}) {w}")
                writer.writerow(row)
            for i in range(2):
                row = [region, "S16", f"S16 Game {i+1}"]
                for b in brackets:
                    w = b[region]["S16"][i]
                    row.append(f"({TEAMS[w]['seed']}) {w}")
                writer.writerow(row)
            row = [region, "E8", f"{region} Champion"]
            for b in brackets:
                w = b[region]["E8"]
                row.append(f"({TEAMS[w]['seed']}) {w}")
            writer.writerow(row)

        for i, (ra, rb) in enumerate([("South", "East"), ("Midwest", "West")]):
            row = ["", "FF", f"Semifinal ({ra} vs {rb})"]
            for b in brackets:
                w = b["FF"][i]
                row.append(f"({TEAMS[w]['seed']}) {w}")
            writer.writerow(row)

        row = ["", "CG", "NATIONAL CHAMPION"]
        for b in brackets:
            c = b["Champion"]
            row.append(f"({TEAMS[c]['seed']}) {c}")
        writer.writerow(row)

    print("Saved: output/all_20_brackets.csv")
